#!/usr/bin/env python3
"""
Allen Brain Cell Atlas Query Tool

Lightweight CLI for querying the Allen Brain Cell Atlas taxonomy data,
brain structure ontology, and cell types database.

Usage:
    python abc_query.py lookup "Sst"
    python abc_query.py lookup "Pvalb" --species human
    python abc_query.py markers "L5 IT"
    python abc_query.py region "CTX"
    python abc_query.py hierarchy class
    python abc_query.py hierarchy subclass --filter GABA
    python abc_query.py search "dopaminergic"
    python abc_query.py structure "hippocampus"
    python abc_query.py specimen --region "VISp" --type spiny
    python abc_query.py specimen --id 485909730

Dependencies: openpyxl (pip install openpyxl)
"""

import argparse
import csv
import json
import sys
import urllib.parse
import urllib.request
from pathlib import Path

# ---------------------------------------------------------------------------
# Data source URLs
# ---------------------------------------------------------------------------

S3_BASE_MOUSE = "https://allen-brain-cell-atlas.s3.us-west-2.amazonaws.com/metadata/WMB-taxonomy/20231215"
S3_BASE_HUMAN = "https://allen-brain-cell-atlas.s3.us-west-2.amazonaws.com/metadata/WHB-taxonomy/20240330"
ALLEN_API = "https://api.brain-map.org/api/v2"

FILES_MOUSE = {
    "pivoted": f"{S3_BASE_MOUSE}/views/cluster_to_cluster_annotation_membership_pivoted.csv",
    "terms_with_counts": f"{S3_BASE_MOUSE}/views/cluster_annotation_term_with_counts.csv",
    "cluster_counts": f"{S3_BASE_MOUSE}/cluster.csv",
    "annotations": f"{S3_BASE_MOUSE}/cl.df_CCN202307220.xlsx",
}

FILES_HUMAN = {
    "terms": f"{S3_BASE_HUMAN}/cluster_annotation_term.csv",
    "term_sets": f"{S3_BASE_HUMAN}/cluster_annotation_term_set.csv",
    "cluster_counts": f"{S3_BASE_HUMAN}/cluster.csv",
    "membership": f"{S3_BASE_HUMAN}/cluster_to_cluster_annotation_membership.csv",
}

CACHE_DIR = Path.home() / ".abc_atlas_cache"

# Common synonyms -> atlas abbreviations
SYNONYMS = {
    "dopaminergic": "Dopa",
    "dopamine": "Dopa",
    "serotonergic": "Sero",
    "serotonin": "Sero",
    "cholinergic": "Chol",
    "glutamatergic": "Glut",
    "glutamate": "Glut",
    "gabaergic": "GABA",
    "inhibitory": "GABA",
    "excitatory": "Glut",
    "parvalbumin": "Pvalb",
    "somatostatin": "Sst",
    "astrocyte": "Astro",
    "astrocytes": "Astro",
    "oligodendrocyte": "Oligo",
    "oligodendrocytes": "Oligo",
    "microglia": "Immune",
    "hippocampus": "HIP",
    "hippocampal": "HIP",
    "cortex": "CTX",
    "cortical": "CTX",
    "thalamus": "TH",
    "thalamic": "TH",
    "hypothalamus": "HY",
    "hypothalamic": "HY",
    "cerebellum": "CB",
    "cerebellar": "CB",
    "midbrain": "MB",
    "hindbrain": "HB",
    "striatum": "STR",
    "striatal": "STR",
    "amygdala": "Amyg",
    "olfactory bulb": "OB",
}


def expand_synonyms(query: str) -> list[str]:
    """Return the query plus any synonym expansions."""
    queries = [query]
    lower = query.lower()
    if lower in SYNONYMS:
        queries.append(SYNONYMS[lower])
    return queries


# ---------------------------------------------------------------------------
# Nomenclature database
# ---------------------------------------------------------------------------

def load_nomenclature() -> dict:
    """Load the nomenclature alias database."""
    nom_path = Path(__file__).parent / "nomenclature.json"
    if not nom_path.exists():
        print(json.dumps({"error": f"nomenclature.json not found at {nom_path}"}))
        sys.exit(1)
    with open(nom_path) as f:
        return json.load(f)


def score_alias_match(query_lower: str, alias_name_lower: str) -> int:
    """Score how well a query matches an alias. Higher = better. 0 = no match."""
    if query_lower == alias_name_lower:
        return 100  # exact
    # Check if query is a substring of alias or vice versa
    if query_lower in alias_name_lower or alias_name_lower in query_lower:
        return 60
    # Check token overlap
    q_tokens = set(query_lower.split())
    a_tokens = set(alias_name_lower.split())
    overlap = q_tokens & a_tokens
    if overlap and len(overlap) >= len(q_tokens) * 0.5:
        return 40 + len(overlap) * 10
    return 0


def cmd_resolve(args):
    """Resolve a cell type name to its canonical Allen taxonomy entry."""
    query = args.query
    query_lower = query.lower()
    region_filter = getattr(args, "region", None)
    nom = load_nomenclature()

    matches = []

    # Search entries
    for entry in nom.get("entries", []):
        canonical = entry["canonical"]
        entry_region = entry.get("region", "")

        # Check canonical name
        if query_lower == canonical.lower():
            matches.append({
                "canonical": canonical,
                "level": entry.get("level"),
                "class": entry.get("class"),
                "region": entry_region,
                "match_type": "canonical_exact",
                "match_quality": "exact",
                "score": 100,
            })
            continue

        # Check aliases
        for alias in entry.get("aliases", []):
            score = score_alias_match(query_lower, alias["name"].lower())
            if score > 0:
                if region_filter and entry_region and region_filter.lower() not in entry_region.lower():
                    continue
                matches.append({
                    "canonical": canonical,
                    "level": entry.get("level"),
                    "class": entry.get("class"),
                    "region": entry_region,
                    "matched_alias": alias["name"],
                    "alias_source": alias.get("source", ""),
                    "match_type": "alias",
                    "match_quality": alias.get("match_quality", "exact"),
                    "note": alias.get("note"),
                    "score": score,
                })

        # Check supertypes
        for st in entry.get("supertypes", []):
            st_canonical = st["canonical"]
            if query_lower == st_canonical.lower():
                matches.append({
                    "canonical": st_canonical,
                    "parent_subclass": canonical,
                    "level": "supertype",
                    "class": entry.get("class"),
                    "region": entry_region,
                    "match_type": "canonical_exact",
                    "match_quality": "exact",
                    "score": 100,
                })
                continue
            for alias in st.get("aliases", []):
                score = score_alias_match(query_lower, alias["name"].lower())
                if score > 0:
                    if region_filter and entry_region and region_filter.lower() not in entry_region.lower():
                        continue
                    matches.append({
                        "canonical": st_canonical,
                        "parent_subclass": canonical,
                        "level": "supertype",
                        "class": entry.get("class"),
                        "region": entry_region,
                        "matched_alias": alias["name"],
                        "alias_source": alias.get("source", ""),
                        "match_type": "alias",
                        "match_quality": alias.get("match_quality", "exact"),
                        "note": st.get("note"),
                        "score": score,
                    })

    # Check ambiguous terms
    ambiguous_hit = None
    for amb in nom.get("ambiguous_terms", []):
        if query_lower == amb["term"].lower() or query_lower in amb["term"].lower():
            if region_filter:
                # Filter candidates by region
                filtered = [c for c in amb["candidates"]
                            if region_filter.lower() in c.get("region", "").lower()]
                if filtered:
                    ambiguous_hit = None  # resolved by region
                    for c in filtered:
                        matches.append({
                            "canonical": c["canonical"],
                            "region": c.get("region"),
                            "description": c.get("description"),
                            "match_type": "disambiguated_by_region",
                            "match_quality": "exact",
                            "score": 80,
                        })
                else:
                    ambiguous_hit = amb
            else:
                ambiguous_hit = amb

    # Sort by score
    matches.sort(key=lambda x: x["score"], reverse=True)

    # Deduplicate by canonical name, keeping highest score
    seen = set()
    unique = []
    for m in matches:
        key = m["canonical"]
        if key not in seen:
            seen.add(key)
            unique.append(m)

    if len(unique) > args.limit:
        unique = unique[:args.limit]

    output = {"query": query, "resolved": len(unique), "results": unique}
    if ambiguous_hit:
        output["ambiguous"] = True
        output["disambiguation"] = {
            "term": ambiguous_hit["term"],
            "hint": ambiguous_hit.get("disambiguation_hint", ""),
            "candidates": ambiguous_hit["candidates"],
        }
    if region_filter:
        output["region_filter"] = region_filter

    print(json.dumps(output, indent=2, default=str))


# ---------------------------------------------------------------------------
# Cache & loading helpers
# ---------------------------------------------------------------------------

def ensure_cached(url: str, filename: str = None) -> Path:
    """Download a file from a URL if not already cached."""
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    if filename is None:
        filename = url.split("/")[-1]
    local_path = CACHE_DIR / filename
    if not local_path.exists():
        print(f"Downloading {filename}...", file=sys.stderr)
        urllib.request.urlretrieve(url, local_path)
    return local_path


def ensure_cached_legacy(name: str) -> Path:
    """Download a mouse taxonomy file by key name."""
    url = FILES_MOUSE[name]
    return ensure_cached(url)


def load_csv(path: Path) -> list[dict]:
    """Load a CSV file as a list of dicts."""
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def load_annotations() -> dict:
    """Load the mouse Excel annotations file. Returns dict of sheet_name -> list of row dicts."""
    try:
        import openpyxl
    except ImportError:
        print("Error: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    path = ensure_cached_legacy("annotations")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        result[sheet_name] = [
            {h: v for h, v in zip(headers, row)} for row in rows[1:]
        ]
    wb.close()
    return result


def load_cluster_counts() -> dict[str, int]:
    """Load mouse cluster alias -> cell count mapping."""
    path = ensure_cached_legacy("cluster_counts")
    rows = load_csv(path)
    return {row["cluster_alias"]: int(row["number_of_cells"]) for row in rows}


def fuzzy_match(query: str, text: str) -> bool:
    """Case-insensitive substring match with synonym expansion."""
    text_lower = text.lower()
    for q in expand_synonyms(query):
        if q.lower() in text_lower:
            return True
    return False


def api_query(url: str) -> dict:
    """Fetch JSON from the Allen Brain Map API."""
    # Split base and query, encode query params properly
    parts = url.split("?", 1)
    if len(parts) == 2:
        encoded_url = parts[0] + "?" + urllib.parse.quote(parts[1], safe="=&[]$,*:'")
    else:
        encoded_url = url
    try:
        req = urllib.request.Request(encoded_url)
        with urllib.request.urlopen(req, timeout=30) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except urllib.error.URLError as e:
        print(json.dumps({"error": f"API request failed: {e}"}))
        sys.exit(1)
    except TimeoutError:
        print(json.dumps({"error": "API request timed out"}))
        sys.exit(1)


# ---------------------------------------------------------------------------
# Human taxonomy helpers
# ---------------------------------------------------------------------------

def load_human_terms() -> list[dict]:
    """Load human brain taxonomy annotation terms with cell counts."""
    path = ensure_cached(FILES_HUMAN["terms"], "whb_cluster_annotation_term.csv")
    return load_csv(path)


def load_human_cluster_counts() -> dict[str, int]:
    """Load human cluster alias -> cell count mapping."""
    path = ensure_cached(FILES_HUMAN["cluster_counts"], "whb_cluster.csv")
    rows = load_csv(path)
    return {row["cluster_alias"]: int(row["number_of_cells"]) for row in rows}


def load_human_membership() -> list[dict]:
    """Load human cluster -> hierarchy membership."""
    path = ensure_cached(FILES_HUMAN["membership"], "whb_cluster_to_cluster_annotation_membership.csv")
    return load_csv(path)


def build_human_hierarchy() -> list[dict]:
    """Build a flat lookup from human membership + terms data.

    Returns one entry per subcluster with its parent cluster and supercluster.
    """
    terms = load_human_terms()
    # Build label -> term info map
    term_map = {}
    for t in terms:
        term_map[t["label"]] = t

    # Build membership: cluster_alias -> list of term labels
    membership = load_human_membership()
    alias_terms = {}
    for row in membership:
        alias = row.get("cluster_alias", "")
        term_label = row.get("cluster_annotation_term_label", "")
        if alias not in alias_terms:
            alias_terms[alias] = {}
        term_info = term_map.get(term_label, {})
        term_set = term_info.get("cluster_annotation_term_set_name", "")
        alias_terms[alias][term_set] = {
            "label": term_label,
            "name": term_info.get("name", ""),
            "cell_count": term_info.get("number_of_cells", ""),
        }

    cluster_counts = load_human_cluster_counts()
    results = []
    for alias, terms_by_level in alias_terms.items():
        entry = {
            "cluster_alias": alias,
            "cell_count": cluster_counts.get(alias),
        }
        for level_name in ("supercluster", "cluster", "subcluster", "neurotransmitter"):
            info = terms_by_level.get(level_name, {})
            entry[level_name] = info.get("name", "")
        results.append(entry)

    return results


# ---------------------------------------------------------------------------
# Commands: Mouse taxonomy (existing)
# ---------------------------------------------------------------------------

def cmd_lookup(args):
    """Look up a cell type by name. Returns hierarchy, markers, and brain region."""
    query = args.query
    species = getattr(args, "species", "mouse")

    if species == "human":
        return cmd_lookup_human(query, args.limit)

    annotations = load_annotations()
    cluster_counts = load_cluster_counts()
    results = []

    # Search cluster-level annotations
    for row in annotations.get("cluster_annotation", []):
        searchable = " ".join(str(v) for v in [
            row.get("cluster_id_label", ""),
            row.get("supertype_label", ""),
            row.get("subclass_label", ""),
            row.get("class_label", ""),
            row.get("anatomical_annotation", ""),
            row.get("nt_type_label", ""),
        ])
        if fuzzy_match(query, searchable):
            cluster_alias = str(row.get("cluster_id", ""))
            cell_count = cluster_counts.get(cluster_alias, None)
            results.append({
                "level": "cluster",
                "cluster_id": row.get("cluster_id_label"),
                "supertype": row.get("supertype_id_label"),
                "subclass": row.get("subclass_id_label"),
                "class": row.get("class_id_label"),
                "neurotransmitter": row.get("nt_type_label"),
                "anatomical_annotation": row.get("anatomical_annotation"),
                "neighborhood": row.get("neighborhood"),
                "marker_genes": row.get("cluster.markers.combo"),
                "merfish_markers": row.get("merfish.markers.combo"),
                "tf_markers": row.get("cluster.TF.markers.combo"),
                "nt_markers": row.get("nt.markers"),
                "cell_count": cell_count,
            })

    # Search supertype-level
    for row in annotations.get("supertype_annotation", []):
        searchable = " ".join(str(v) for v in [
            row.get("supertype_id_label", ""),
            row.get("supertype_label", ""),
            row.get("subclass_label", ""),
            row.get("class_label", ""),
        ])
        if fuzzy_match(query, searchable):
            results.append({
                "level": "supertype",
                "supertype": row.get("supertype_id_label"),
                "subclass": row.get("subclass_id_label"),
                "class": row.get("class_id_label"),
                "marker_genes": row.get("supertype.markers.combo"),
                "markers_within_subclass": row.get("supertype.markers.combo (within subclass)"),
            })

    # Search subclass-level
    for row in annotations.get("subclass_annotation", []):
        searchable = " ".join(str(v) for v in [
            row.get("subclass_id_label", ""),
            row.get("subclass_label", ""),
            row.get("class_label", ""),
            row.get("nt_type_label", ""),
        ])
        if fuzzy_match(query, searchable):
            results.append({
                "level": "subclass",
                "subclass": row.get("subclass_id_label"),
                "class": row.get("class_id_label"),
                "neurotransmitter": row.get("nt_type_label"),
                "neighborhood": row.get("neighborhood"),
                "marker_genes": row.get("subclass.markers.combo"),
                "tf_markers": row.get("subclass.tf.markers.combo"),
            })

    # Search class-level
    for row in annotations.get("class_annotation", []):
        searchable = " ".join(str(v) for v in [
            row.get("class_id_label", ""),
            row.get("class_label", ""),
            row.get("neighborhood", ""),
        ])
        if fuzzy_match(query, searchable):
            results.append({
                "level": "class",
                "class": row.get("class_id_label"),
                "neighborhood": row.get("neighborhood"),
            })

    if len(results) > args.limit:
        results = results[:args.limit]

    output = {"query": query, "species": "mouse", "total_matches": len(results), "results": results}
    print(json.dumps(output, indent=2, default=str))


def cmd_lookup_human(query: str, limit: int):
    """Look up a cell type in the human brain taxonomy."""
    terms = load_human_terms()
    results = []

    for term in terms:
        searchable = f"{term.get('name', '')} {term.get('label', '')} {term.get('description', '')}"
        if fuzzy_match(query, searchable):
            cell_count = term.get("number_of_cells", "")
            try:
                cell_count = int(cell_count)
            except (ValueError, TypeError):
                pass
            results.append({
                "level": term.get("cluster_annotation_term_set_name", ""),
                "name": term.get("name", ""),
                "label": term.get("label", ""),
                "parent": term.get("parent_term_label", ""),
                "cell_count": cell_count,
                "description": term.get("description", ""),
            })

    if len(results) > limit:
        results = results[:limit]

    output = {"query": query, "species": "human", "total_matches": len(results), "results": results}
    print(json.dumps(output, indent=2, default=str))


def cmd_markers(args):
    """Return marker genes for a cell type at any hierarchy level."""
    query = args.query
    annotations = load_annotations()
    results = []

    # Check subclass level first (most useful)
    for row in annotations.get("subclass_annotation", []):
        searchable = f"{row.get('subclass_label', '')} {row.get('subclass_id_label', '')}"
        if fuzzy_match(query, searchable):
            results.append({
                "level": "subclass",
                "name": row.get("subclass_id_label"),
                "neurotransmitter": row.get("nt_type_label"),
                "marker_genes": row.get("subclass.markers.combo"),
                "tf_markers": row.get("subclass.tf.markers.combo"),
            })

    # Supertype level
    for row in annotations.get("supertype_annotation", []):
        searchable = f"{row.get('supertype_label', '')} {row.get('supertype_id_label', '')}"
        if fuzzy_match(query, searchable):
            results.append({
                "level": "supertype",
                "name": row.get("supertype_id_label"),
                "subclass": row.get("subclass_id_label"),
                "marker_genes": row.get("supertype.markers.combo"),
                "markers_within_subclass": row.get("supertype.markers.combo (within subclass)"),
            })

    # Cluster level
    for row in annotations.get("cluster_annotation", []):
        searchable = f"{row.get('cluster_id_label', '')} {row.get('supertype_label', '')}"
        if fuzzy_match(query, searchable):
            results.append({
                "level": "cluster",
                "name": row.get("cluster_id_label"),
                "supertype": row.get("supertype_id_label"),
                "marker_genes": row.get("cluster.markers.combo"),
                "merfish_markers": row.get("merfish.markers.combo"),
                "tf_markers": row.get("cluster.TF.markers.combo"),
            })

    if len(results) > args.limit:
        results = results[:args.limit]

    output = {"query": query, "total_matches": len(results), "results": results}
    print(json.dumps(output, indent=2, default=str))


def cmd_gene(args):
    """Find which cell types express a given gene (reverse marker lookup)."""
    query = args.query.lower()
    annotations = load_annotations()
    results = []

    # Search subclass markers
    for row in annotations.get("subclass_annotation", []):
        markers = str(row.get("subclass.markers.combo", ""))
        tf_markers = str(row.get("subclass.tf.markers.combo", ""))
        all_markers = f"{markers},{tf_markers}"
        gene_list = [g.strip().lower() for g in all_markers.split(",") if g.strip()]
        if query in gene_list:
            results.append({
                "level": "subclass",
                "name": row.get("subclass_id_label"),
                "class": row.get("class_id_label"),
                "neurotransmitter": row.get("nt_type_label"),
                "marker_type": "subclass_marker" if query in markers.lower() else "tf_marker",
            })

    # Search supertype markers
    for row in annotations.get("supertype_annotation", []):
        markers = str(row.get("supertype.markers.combo", ""))
        within = str(row.get("supertype.markers.combo (within subclass)", ""))
        all_markers = f"{markers},{within}"
        gene_list = [g.strip().lower() for g in all_markers.split(",") if g.strip()]
        if query in gene_list:
            results.append({
                "level": "supertype",
                "name": row.get("supertype_id_label"),
                "subclass": row.get("subclass_id_label"),
                "class": row.get("class_id_label"),
                "marker_type": "supertype_marker" if query in markers.lower() else "within_subclass_marker",
            })

    # Search cluster markers
    for row in annotations.get("cluster_annotation", []):
        cluster_m = str(row.get("cluster.markers.combo", ""))
        merfish_m = str(row.get("merfish.markers.combo", ""))
        tf_m = str(row.get("cluster.TF.markers.combo", ""))
        nt_m = str(row.get("nt.markers", ""))
        # nt.markers has format "Gene:value,Gene:value" — extract gene names
        nt_genes = [g.split(":")[0].strip().lower() for g in nt_m.split(",") if g.strip()]
        all_genes = set()
        for m in (cluster_m, merfish_m, tf_m):
            all_genes.update(g.strip().lower() for g in m.split(",") if g.strip())
        all_genes.update(nt_genes)
        if query in all_genes:
            marker_types = []
            if query in [g.strip().lower() for g in cluster_m.split(",")]:
                marker_types.append("cluster_marker")
            if query in [g.strip().lower() for g in merfish_m.split(",")]:
                marker_types.append("merfish_marker")
            if query in [g.strip().lower() for g in tf_m.split(",")]:
                marker_types.append("tf_marker")
            if query in nt_genes:
                marker_types.append("nt_marker")
            results.append({
                "level": "cluster",
                "name": row.get("cluster_id_label"),
                "supertype": row.get("supertype_id_label"),
                "subclass": row.get("subclass_id_label"),
                "class": row.get("class_id_label"),
                "marker_types": marker_types,
            })

    if len(results) > args.limit:
        results = results[:args.limit]

    output = {"query": args.query, "total_matches": len(results), "results": results}
    print(json.dumps(output, indent=2, default=str))


def cmd_region(args):
    """List cell types found in a brain region."""
    query = args.query
    annotations = load_annotations()
    cluster_counts = load_cluster_counts()

    matches = []
    for row in annotations.get("cluster_annotation", []):
        anat = str(row.get("anatomical_annotation", ""))
        ccf_broad = str(row.get("CCF_broad.freq", ""))
        ccf_acr = str(row.get("CCF_acronym.freq", ""))
        if fuzzy_match(query, anat) or fuzzy_match(query, ccf_broad) or fuzzy_match(query, ccf_acr):
            matches.append(row)

    # Aggregate by subclass
    subclass_map = {}
    for row in matches:
        subclass = row.get("subclass_id_label", "unknown")
        if subclass not in subclass_map:
            subclass_map[subclass] = {
                "subclass": subclass,
                "class": row.get("class_id_label"),
                "neurotransmitter": row.get("nt_type_label"),
                "cluster_count": 0,
                "total_cells": 0,
                "anatomical_annotations": set(),
            }
        subclass_map[subclass]["cluster_count"] += 1
        alias = str(row.get("cluster_id", ""))
        subclass_map[subclass]["total_cells"] += cluster_counts.get(alias, 0)
        anat = row.get("anatomical_annotation")
        if anat:
            subclass_map[subclass]["anatomical_annotations"].add(str(anat))

    results = sorted(subclass_map.values(), key=lambda x: x["total_cells"], reverse=True)
    for r in results:
        r["anatomical_annotations"] = sorted(r["anatomical_annotations"])

    if len(results) > args.limit:
        results = results[:args.limit]

    output = {"query": query, "region_matches": len(results), "results": results}
    print(json.dumps(output, indent=2, default=str))


def cmd_hierarchy(args):
    """List all entries at a given taxonomy level."""
    level = args.level
    filter_term = args.filter
    species = getattr(args, "species", "mouse")

    if species == "human":
        return cmd_hierarchy_human(level, filter_term, args.limit)

    annotations = load_annotations()

    sheet_map = {
        "class": "class_annotation",
        "subclass": "subclass_annotation",
        "supertype": "supertype_annotation",
        "cluster": "cluster_annotation",
    }

    if level not in sheet_map:
        print(f"Error: level must be one of: {', '.join(sheet_map.keys())}", file=sys.stderr)
        sys.exit(1)

    sheet = sheet_map[level]
    rows = annotations.get(sheet, [])

    label_key = f"{level}_id_label"
    results = []
    for row in rows:
        label = str(row.get(label_key, row.get(f"{level}_label", "")))
        if filter_term and not fuzzy_match(filter_term, label):
            class_label = str(row.get("class_label", row.get("class_id_label", "")))
            nt_label = str(row.get("nt_type_label", ""))
            neighborhood = str(row.get("neighborhood", ""))
            if not any(fuzzy_match(filter_term, f) for f in [class_label, nt_label, neighborhood]):
                continue

        entry = {"name": label}
        if level == "subclass":
            entry["class"] = row.get("class_id_label")
            entry["neurotransmitter"] = row.get("nt_type_label")
            entry["neighborhood"] = row.get("neighborhood")
            entry["marker_genes"] = row.get("subclass.markers.combo")
        elif level == "supertype":
            entry["subclass"] = row.get("subclass_id_label")
            entry["class"] = row.get("class_id_label")
            entry["marker_genes"] = row.get("supertype.markers.combo")
        elif level == "class":
            entry["neighborhood"] = row.get("neighborhood")
        elif level == "cluster":
            entry["supertype"] = row.get("supertype_id_label")
            entry["subclass"] = row.get("subclass_id_label")
            entry["class"] = row.get("class_id_label")
            entry["anatomical_annotation"] = row.get("anatomical_annotation")

        results.append(entry)

    if len(results) > args.limit:
        results = results[:args.limit]

    output = {"level": level, "species": "mouse", "filter": filter_term, "total": len(results), "results": results}
    print(json.dumps(output, indent=2, default=str))


def cmd_hierarchy_human(level: str, filter_term: str, limit: int):
    """List human taxonomy entries at a given level."""
    # Human levels: supercluster, cluster, subcluster, neurotransmitter
    valid_levels = {"supercluster", "cluster", "subcluster", "neurotransmitter"}
    if level not in valid_levels:
        print(f"Error: for human, level must be one of: {', '.join(sorted(valid_levels))}", file=sys.stderr)
        print("(Mouse levels: class, subclass, supertype, cluster)", file=sys.stderr)
        sys.exit(1)

    terms = load_human_terms()
    results = []
    for term in terms:
        if term.get("cluster_annotation_term_set_name", "") != level:
            continue
        name = term.get("name", "")
        if filter_term and not fuzzy_match(filter_term, name):
            desc = str(term.get("description", ""))
            if not fuzzy_match(filter_term, desc):
                continue
        cell_count = term.get("number_of_cells", "")
        try:
            cell_count = int(cell_count)
        except (ValueError, TypeError):
            pass
        results.append({
            "name": name,
            "label": term.get("label", ""),
            "cell_count": cell_count,
            "parent": term.get("parent_term_label", ""),
        })

    if len(results) > limit:
        results = results[:limit]

    output = {"level": level, "species": "human", "filter": filter_term, "total": len(results), "results": results}
    print(json.dumps(output, indent=2, default=str))


def cmd_search(args):
    """Free-text search across all cell type names and annotations."""
    query = args.query
    species = getattr(args, "species", "mouse")
    results = []

    if species == "human" or species == "both":
        terms = load_human_terms()
        for term in terms:
            searchable = " ".join(str(v) for v in term.values() if v)
            if fuzzy_match(query, searchable):
                cell_count = term.get("number_of_cells", "")
                try:
                    cell_count = int(cell_count)
                except (ValueError, TypeError):
                    pass
                results.append({
                    "species": "human",
                    "level": term.get("cluster_annotation_term_set_name", ""),
                    "name": term.get("name", ""),
                    "cell_count": cell_count,
                    "description": term.get("description", ""),
                })

    if species == "mouse" or species == "both":
        annotations = load_annotations()
        for sheet_name, rows in annotations.items():
            if sheet_name in ("MERFISH_gene_panel", "Column header explanation"):
                continue
            for row in rows:
                all_text = " ".join(str(v) for v in row.values() if v is not None)
                if fuzzy_match(query, all_text):
                    entry = {"species": "mouse", "source": sheet_name}
                    for key in ("cluster_id_label", "supertype_id_label", "supertype_label",
                                "subclass_id_label", "subclass_label", "class_id_label",
                                "class_label", "anatomical_annotation", "nt_type_label",
                                "neighborhood"):
                        if key in row and row[key] is not None:
                            entry[key.replace("_id_label", "").replace("_label", "")] = row[key]
                    results.append(entry)

    # Deduplicate
    seen = set()
    unique = []
    for r in results:
        key = json.dumps(r, sort_keys=True, default=str)
        if key not in seen:
            seen.add(key)
            unique.append(r)

    if len(unique) > args.limit:
        unique = unique[:args.limit]

    output = {"query": query, "species": species, "total_matches": len(unique), "results": unique}
    print(json.dumps(output, indent=2, default=str))


# ---------------------------------------------------------------------------
# Commands: Brain structure ontology
# ---------------------------------------------------------------------------

def load_structure_ontology() -> dict:
    """Load the Allen Mouse Brain structure ontology. Cached as JSON."""
    cache_path = CACHE_DIR / "structure_ontology.json"
    if not cache_path.exists():
        CACHE_DIR.mkdir(parents=True, exist_ok=True)
        print("Downloading brain structure ontology...", file=sys.stderr)
        url = f"{ALLEN_API}/structure_graph_download/1.json"
        data = api_query(url)
        with open(cache_path, "w") as f:
            json.dump(data, f)
        return data
    with open(cache_path) as f:
        return json.load(f)


def flatten_structures(node: dict, parent_name: str = None, depth: int = 0) -> list[dict]:
    """Recursively flatten the structure tree into a list."""
    entry = {
        "id": node.get("id"),
        "acronym": node.get("acronym"),
        "name": node.get("name"),
        "parent": parent_name,
        "depth": depth,
        "color": node.get("color_hex_triplet"),
    }
    result = [entry]
    for child in node.get("children", []):
        result.extend(flatten_structures(child, node.get("name"), depth + 1))
    return result


def cmd_structure(args):
    """Look up brain structures by name or acronym."""
    query = args.query
    data = load_structure_ontology()
    root = data.get("msg", [{}])[0]
    all_structures = flatten_structures(root)

    results = []
    for s in all_structures:
        searchable = f"{s['name']} {s['acronym']}"
        if fuzzy_match(query, searchable):
            results.append(s)

    if args.children and len(results) == 1:
        # If exactly one match, also show its direct children
        target_name = results[0]["name"]
        children = [s for s in all_structures if s["parent"] == target_name]
        results[0]["children"] = [
            {"acronym": c["acronym"], "name": c["name"], "id": c["id"]}
            for c in children
        ]

    if len(results) > args.limit:
        results = results[:args.limit]

    output = {"query": query, "total_matches": len(results), "results": results}
    print(json.dumps(output, indent=2, default=str))


def cmd_structure_path(args):
    """Show the full path from root to a brain structure."""
    query = args.query
    data = load_structure_ontology()
    root = data.get("msg", [{}])[0]
    all_structures = flatten_structures(root)

    # Find the target structure
    target = None
    for s in all_structures:
        if s["acronym"].lower() == query.lower() or s["name"].lower() == query.lower():
            target = s
            break
    if not target:
        # Try fuzzy
        for s in all_structures:
            if fuzzy_match(query, f"{s['name']} {s['acronym']}"):
                target = s
                break

    if not target:
        print(json.dumps({"query": query, "error": "Structure not found"}))
        return

    # Walk up the parent chain
    path = [{"acronym": target["acronym"], "name": target["name"]}]
    current_parent = target["parent"]
    while current_parent:
        parent_node = next((s for s in all_structures if s["name"] == current_parent), None)
        if not parent_node:
            break
        path.append({"acronym": parent_node["acronym"], "name": parent_node["name"]})
        current_parent = parent_node["parent"]

    path.reverse()

    # Also get siblings (structures at same level with same parent)
    siblings = [
        {"acronym": s["acronym"], "name": s["name"]}
        for s in all_structures
        if s["parent"] == target["parent"] and s["name"] != target["name"]
    ]

    output = {
        "query": query,
        "structure": {"acronym": target["acronym"], "name": target["name"], "id": target["id"]},
        "path_from_root": path,
        "siblings": siblings[:20],
    }
    print(json.dumps(output, indent=2, default=str))


# ---------------------------------------------------------------------------
# Commands: Cell Types Database (electrophysiology & morphology)
# ---------------------------------------------------------------------------

def cmd_specimen(args):
    """Query the Allen Cell Types Database for specimens with ephys/morphology data."""
    if args.id:
        # Fetch a specific specimen
        url = (
            f"{ALLEN_API}/data/query.json?"
            f"criteria=model::ApiCellTypesSpecimenDetail,"
            f"rma::criteria,[specimen__id$eq{args.id}],"
            f"rma::options[num_rows$eqall]"
        )
        data = api_query(url)
        specimens = data.get("msg", [])
        if not specimens:
            print(json.dumps({"error": f"No specimen found with id {args.id}"}))
            return
        results = [format_specimen(s) for s in specimens]
        print(json.dumps({"specimen_id": args.id, "results": results}, indent=2, default=str))
        return

    # Build filter query
    filters = []
    if args.species:
        species_map = {"mouse": "Mus musculus", "human": "Homo Sapiens"}
        species_name = species_map.get(args.species.lower(), args.species)
        filters.append(f"[donor__species$il'{species_name}']")
    if args.region:
        filters.append(f"[structure__acronym$il'*{args.region}*']")
    if args.type:
        filters.append(f"[tag__dendrite_type$eq'{args.type}']")
    if args.layer:
        filters.append(f"[structure__layer$eq'{args.layer}']")
    if args.has_morphology:
        filters.append("[nr__reconstruction_type$nenull]")
    if args.has_model:
        filters.append("[m__glif$gt0]")

    filter_str = ",".join(filters)
    if filter_str:
        url = (
            f"{ALLEN_API}/data/query.json?"
            f"criteria=model::ApiCellTypesSpecimenDetail,"
            f"rma::criteria,{filter_str},"
            f"rma::options[num_rows$eq{args.limit}]"
        )
    else:
        url = (
            f"{ALLEN_API}/data/query.json?"
            f"criteria=model::ApiCellTypesSpecimenDetail,"
            f"rma::options[num_rows$eq{args.limit}]"
        )
    data = api_query(url)
    specimens = data.get("msg", [])
    total = data.get("total_rows", len(specimens))

    results = [format_specimen(s) for s in specimens]
    output = {
        "total_available": total,
        "returned": len(results),
        "filters": {
            "species": args.species,
            "region": args.region,
            "dendrite_type": args.type,
            "layer": args.layer,
        },
        "results": results,
    }
    print(json.dumps(output, indent=2, default=str))


def format_specimen(s: dict) -> dict:
    """Format a specimen record into a clean dict."""
    specimen_id = s.get("specimen__id")
    result = {
        "specimen_id": specimen_id,
        "name": s.get("specimen__name"),
        "species": s.get("donor__species"),
        "sex": s.get("donor__sex"),
        "age": s.get("donor__age"),
        "disease_state": s.get("donor__disease_state"),
        "brain_region": s.get("structure__name"),
        "region_acronym": s.get("structure__acronym"),
        "layer": s.get("structure__layer"),
        "hemisphere": s.get("specimen__hemisphere"),
        "dendrite_type": s.get("tag__dendrite_type"),
        "apical": s.get("tag__apical"),
    }
    if specimen_id:
        result["web_url"] = f"https://celltypes.brain-map.org/experiment/electrophysiology/{specimen_id}"

    # Electrophysiology features
    ephys = {}
    for key in ("ef__vrest", "ef__ri", "ef__tau", "ef__avg_firing_rate",
                "ef__adaptation", "ef__f_i_curve_slope",
                "ef__threshold_i_long_square", "ef__upstroke_downstroke_ratio_long_square",
                "ef__fast_trough_v_long_square", "ef__peak_t_ramp"):
        val = s.get(key)
        if val is not None:
            ephys[key.replace("ef__", "")] = val
    if ephys:
        result["electrophysiology"] = ephys

    # Morphology features
    morph = {}
    for key in ("nr__reconstruction_type", "nr__number_bifurcations",
                "nr__number_stems", "nr__max_euclidean_distance",
                "nr__average_contraction", "nr__average_parent_daughter_ratio"):
        val = s.get(key)
        if val is not None:
            morph[key.replace("nr__", "")] = val
    if morph:
        result["morphology"] = morph

    # Model availability
    models = {}
    for key in ("m__glif", "m__biophys", "m__biophys_perisomatic", "m__biophys_all_active"):
        val = s.get(key)
        if val and val > 0:
            models[key.replace("m__", "")] = val
    if models:
        result["models_available"] = models

    # Coordinates
    if s.get("csl__x") is not None:
        result["ccf_coordinates"] = {
            "x": s.get("csl__x"),
            "y": s.get("csl__y"),
            "z": s.get("csl__z"),
        }

    return result


# ---------------------------------------------------------------------------
# Main CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Query the Allen Brain Cell Atlas, structure ontology, and cell types database",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s lookup "Sst"                    Look up Sst cell types (mouse)
  %(prog)s lookup "Pvalb" --species human   Look up in human taxonomy
  %(prog)s markers "L5 IT"                  Get marker genes for L5 IT neurons
  %(prog)s gene "Gad2"                     Which cell types express Gad2?
  %(prog)s region "hippocampus"             Cell types in hippocampus
  %(prog)s hierarchy class                  List all 34 mouse classes
  %(prog)s hierarchy supercluster -s human  List human superclusters
  %(prog)s search "dopamine"                Free-text search
  %(prog)s search "Pvalb" -s both           Search both species
  %(prog)s structure "hippocampus"           Brain structure ontology lookup
  %(prog)s structure-path "CA1"             Full path from root to CA1
  %(prog)s resolve "fast-spiking interneuron" Resolve to Allen taxonomy name
  %(prog)s resolve "basket cell" -r cortex  Disambiguate by region
  %(prog)s specimen --region VISp           Specimens from primary visual cortex
  %(prog)s specimen --id 485909730          Specific specimen details
        """,
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    # -- lookup --
    p_lookup = subparsers.add_parser("lookup", help="Look up a cell type by name")
    p_lookup.add_argument("query", help="Cell type name to search for")
    p_lookup.add_argument("--species", "-s", default="mouse", choices=["mouse", "human"],
                          help="Species (default: mouse)")
    p_lookup.add_argument("--limit", "-l", type=int, default=20, help="Max results (default: 20)")
    p_lookup.set_defaults(func=cmd_lookup)

    # -- markers --
    p_markers = subparsers.add_parser("markers", help="Get marker genes for a cell type (mouse only)")
    p_markers.add_argument("query", help="Cell type name")
    p_markers.add_argument("--limit", "-l", type=int, default=20, help="Max results (default: 20)")
    p_markers.set_defaults(func=cmd_markers)

    # -- gene --
    p_gene = subparsers.add_parser("gene", help="Find cell types that express a gene (mouse only)")
    p_gene.add_argument("query", help="Gene symbol (e.g. Gad2, Slc17a7, Pvalb)")
    p_gene.add_argument("--limit", "-l", type=int, default=30, help="Max results (default: 30)")
    p_gene.set_defaults(func=cmd_gene)

    # -- region --
    p_region = subparsers.add_parser("region", help="List cell types in a brain region (mouse)")
    p_region.add_argument("query", help="Brain region name or CCF acronym")
    p_region.add_argument("--limit", "-l", type=int, default=50, help="Max results (default: 50)")
    p_region.set_defaults(func=cmd_region)

    # -- hierarchy --
    p_hier = subparsers.add_parser("hierarchy", help="List taxonomy entries at a level")
    p_hier.add_argument("level",
                        help="Mouse: class|subclass|supertype|cluster. Human: supercluster|cluster|subcluster|neurotransmitter")
    p_hier.add_argument("--species", "-s", default="mouse", choices=["mouse", "human"],
                        help="Species (default: mouse)")
    p_hier.add_argument("--filter", "-f", default=None, help="Filter by keyword")
    p_hier.add_argument("--limit", "-l", type=int, default=100, help="Max results (default: 100)")
    p_hier.set_defaults(func=cmd_hierarchy)

    # -- search --
    p_search = subparsers.add_parser("search", help="Free-text search across all annotations")
    p_search.add_argument("query", help="Search term")
    p_search.add_argument("--species", "-s", default="mouse", choices=["mouse", "human", "both"],
                          help="Species to search (default: mouse)")
    p_search.add_argument("--limit", "-l", type=int, default=30, help="Max results (default: 30)")
    p_search.set_defaults(func=cmd_search)

    # -- structure --
    p_struct = subparsers.add_parser("structure", help="Look up brain structures by name/acronym")
    p_struct.add_argument("query", help="Structure name or acronym (e.g. 'CA1', 'hippocampus')")
    p_struct.add_argument("--children", "-c", action="store_true",
                          help="Show children of matched structure (when single match)")
    p_struct.add_argument("--limit", "-l", type=int, default=20, help="Max results (default: 20)")
    p_struct.set_defaults(func=cmd_structure)

    # -- structure-path --
    p_spath = subparsers.add_parser("structure-path", help="Show full path from root to a structure")
    p_spath.add_argument("query", help="Structure name or acronym")
    p_spath.set_defaults(func=cmd_structure_path)

    # -- resolve --
    p_resolve = subparsers.add_parser("resolve",
                                       help="Resolve a cell type name to its canonical Allen taxonomy entry")
    p_resolve.add_argument("query", help="Cell type name from any naming convention")
    p_resolve.add_argument("--region", "-r", default=None,
                           help="Brain region context to disambiguate (e.g. cortex, cerebellum, hippocampus)")
    p_resolve.add_argument("--limit", "-l", type=int, default=10, help="Max results (default: 10)")
    p_resolve.set_defaults(func=cmd_resolve)

    # -- specimen --
    p_spec = subparsers.add_parser("specimen",
                                   help="Query Cell Types Database for specimens (ephys/morphology)")
    p_spec.add_argument("--id", type=int, default=None, help="Specific specimen ID")
    p_spec.add_argument("--species", default=None, help="Filter by species (mouse/human)")
    p_spec.add_argument("--region", default=None, help="Filter by brain region acronym (e.g. VISp)")
    p_spec.add_argument("--type", default=None, choices=["spiny", "aspiny", "sparsely spiny"],
                        help="Filter by dendrite type")
    p_spec.add_argument("--layer", default=None, help="Filter by cortical layer (e.g. 4, 5)")
    p_spec.add_argument("--has-morphology", action="store_true",
                        help="Only specimens with morphological reconstructions")
    p_spec.add_argument("--has-model", action="store_true",
                        help="Only specimens with GLIF models")
    p_spec.add_argument("--limit", "-l", type=int, default=20, help="Max results (default: 20)")
    p_spec.set_defaults(func=cmd_specimen)

    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
